from __future__ import annotations

from datetime import date
from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
import pytest

import app
from src.krx_openapi import KrxListedSecurityDaily


def test_krx_listed_frame_and_excel_include_requested_columns() -> None:
    trading_date = date(2026, 8, 11)
    frame = app._krx_listed_daily_frame(
        [
            KrxListedSecurityDaily(
                trade_date=trading_date,
                standard_code="KR7005930003",
                short_code="005930",
                stock_name="삼성전자",
                market="KOSPI",
                stock_type="보통주",
                security_type="주권",
                listed_shares=5_969_782_550,
                listing_date=date(1975, 6, 11),
                cumulative_volume=12_345_678,
                cumulative_amount=876_543_210_000,
                is_kospi200=True,
            )
        ]
    )

    assert frame.columns.tolist() == [
        "표준코드",
        "단축코드",
        "종목명",
        "시장구분",
        "주식종류",
        "증권구분",
        "상장주식수",
        "상장일",
        "KRX 거래량",
        "KRX 거래대금",
        "K200",
        "Q150",
    ]
    assert frame.loc[0, "K200"] == "Y"
    assert frame.loc[0, "Q150"] == "-"

    workbook = load_workbook(
        BytesIO(app._krx_listed_excel_bytes(frame, trading_date)),
    )
    worksheet = workbook[f"KRX_{trading_date:%Y%m%d}"]
    assert [cell.value for cell in next(worksheet.iter_rows(max_row=1))] == list(
        frame.columns
    )
    assert worksheet.freeze_panes == "A2"


def test_krx_listed_statistics_follow_filtered_rows() -> None:
    frame = pd.DataFrame(
        [
            {"증권구분": "주권", "주식종류": "보통주"},
            {"증권구분": "주권", "주식종류": "보통주"},
            {"증권구분": "외국주권", "주식종류": "우선주"},
        ]
    )

    security_counts = app._krx_listed_count_frame(frame, "증권구분")
    stock_type_counts = app._krx_listed_count_frame(frame, "주식종류")

    assert security_counts.to_dict("records") == [
        {"증권구분": "주권", "종목수": 2, "비중": 2 / 3},
        {"증권구분": "외국주권", "종목수": 1, "비중": 1 / 3},
    ]
    assert stock_type_counts.to_dict("records") == [
        {"주식종류": "보통주", "종목수": 2, "비중": 2 / 3},
        {"주식종류": "우선주", "종목수": 1, "비중": 1 / 3},
    ]

    workbook = load_workbook(
        BytesIO(
            app._krx_listed_excel_bytes(
                frame,
                date(2026, 8, 12),
                security_counts=security_counts,
                stock_type_counts=stock_type_counts,
            )
        )
    )
    assert workbook.sheetnames == [
        "KRX_20260812",
        "증권구분별 종목수",
        "주식종류별 종목수",
    ]


def test_excel_workbook_formats_rates_and_kind_links() -> None:
    frame = pd.DataFrame(
        [
            {
                "종목코드": "005930",
                "거래량기여도": 0.01234,
                "KIND 공시": (
                    "https://kind.krx.co.kr/common/disclsviewer.do?acptno=1"
                    "#kind-title=투자경고종목지정"
                ),
            }
        ]
    )

    workbook = load_workbook(BytesIO(app._excel_workbook_bytes({"조회 결과": frame})))
    worksheet = workbook["조회 결과"]

    assert worksheet["A2"].value == "005930"
    assert worksheet["B2"].number_format == "0.00%"
    assert worksheet["C2"].value == "투자경고종목지정"
    assert worksheet["C2"].hyperlink.target.startswith("https://kind.krx.co.kr/")


def test_contribution_uses_market_krx_denominator() -> None:
    contributions = [
        app._safe_ratio(100, 1_000),
        app._safe_ratio(50, 1_000),
    ]

    assert contributions == [0.1, 0.05]
    assert sum(item for item in contributions if item is not None) == pytest.approx(0.15)
    assert app._safe_ratio(1, 0) is None


def test_dashboard_contribution_columns_follow_each_stock_ratio() -> None:
    frame = app._comparison_display_frame(
        [
            {
                "종목코드": "005930",
                "종목명": "삼성전자",
                "nxt_current_price": 80_000,
                "change_rate": 0.01,
                "disparity_rate": 0.001,
                "nxt_volume": 100,
                "volume_ratio": 0.2,
                "nxt_amount": 8_000_000,
                "amount_ratio": 0.25,
                "market_cap": 400_000_000_000_000,
                "krx_current_price": 79_900,
                "krx_volume": 500,
                "krx_amount": 32_000_000,
            }
        ],
        {
            "krx_volume": 10_000,
            "krx_amount": 1_000_000_000,
        },
        {"005930": "전체"},
        {"005930": ""},
    )

    assert frame.columns.get_loc("거래량기여도") == frame.columns.get_loc("거래량비율") + 1
    assert frame.columns.get_loc("거래대금기여도") == frame.columns.get_loc("거래대금비율") + 1
    assert frame.loc[0, "거래량기여도"] == 0.01
    assert frame.loc[0, "거래대금기여도"] == 0.008


def test_limit_proximity_excel_frame_uses_visible_labels_and_order() -> None:
    frame = pd.DataFrame(
        [
            {
                "종목코드": "000002",
                "종목명": "하한근접",
                "상장시장": "KOSDAQ",
                "기준가격": 10_000,
                "상한가": 13_000,
                "하한가": 7_000,
                "시가": 7_020,
                "최근접가격": 7_010,
                "근접구분": "하한가",
                "근접시점": "09:10",
                "잔여틱": 1,
                "종가": 7_100,
            },
            {
                "종목코드": "000001",
                "종목명": "상한도달",
                "상장시장": "KOSPI",
                "기준가격": 10_000,
                "상한가": 13_000,
                "하한가": 7_000,
                "시가": 13_000,
                "최근접가격": 13_000,
                "근접구분": "상한가",
                "근접시점": "시가",
                "잔여틱": 0,
                "종가": 12_900,
            },
        ]
    )

    export = app._nxt_limit_proximity_export_frame(frame)

    assert export["종목코드"].tolist() == ["000001", "000002"]
    assert export["상·하한가 근접 구분"].tolist() == ["상한가", "하한가 근접"]
    assert export["도달시점"].tolist() == ["시가", "09:10"]
