from __future__ import annotations

import argparse
import sys
from dataclasses import asdict
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.config import NXT_LAUNCH_DATE
from src.historical_market import HistoricalMarketStore
from src.nxt_change_store import NxtChangeStore
from src.nxt_change_context import contextual_change_reason, source_context_for_change
from src.nxt_eligibility import classify_nxt_change


SUMMARY_COLUMNS = {
    "trade_date": "일자",
    "target_stock_count": "매매체결대상종목수",
    "tradable_stock_count": "거래가능종목수",
    "unavailable_stock_count": "거래불가종목수",
    "target_kospi_count": "매매체결대상_KOSPI종목수",
    "target_kosdaq_count": "매매체결대상_KOSDAQ종목수",
    "tradable_kospi_count": "거래가능_KOSPI종목수",
    "tradable_kosdaq_count": "거래가능_KOSDAQ종목수",
    "inclusion_stock_count": "편입종목수",
    "exclusion_stock_count": "편출종목수",
    "restriction_start_stock_count": "거래불가지정종목수",
    "restriction_end_stock_count": "거래불가해제종목수",
}


def _date_value(raw: str) -> date:
    return date.fromisoformat(raw)


def _summary_frame(store: HistoricalMarketStore, start: date, end: date) -> pd.DataFrame:
    frame = pd.DataFrame(
        [asdict(item) for item in store.list_nxt_eligibility_summaries(start, end)]
    ).rename(columns=SUMMARY_COLUMNS)
    if not frame.empty:
        frame["일자"] = pd.to_datetime(frame["일자"])
    return frame


def _reason_frame(store: HistoricalMarketStore, start: date, end: date) -> pd.DataFrame:
    frame = pd.DataFrame(
        [asdict(item) for item in store.list_nxt_eligibility_reason_counts(start, end)]
    ).rename(
        columns={
            "trade_date": "일자",
            "reason_group": "구분",
            "reason": "사유",
            "stock_count": "종목수",
        }
    )
    if not frame.empty:
        frame["일자"] = pd.to_datetime(frame["일자"])
    return frame


def _wide_summary(summary: pd.DataFrame, reasons: pd.DataFrame) -> pd.DataFrame:
    if summary.empty or reasons.empty:
        return summary
    pivot = reasons.pivot_table(
        index="일자",
        columns=["구분", "사유"],
        values="종목수",
        aggfunc="sum",
        fill_value=0,
    )
    group_order = {
        "거래불가사유": 0,
        "편출": 1,
        "편입": 2,
        "거래불가": 3,
        "거래불가 해제": 4,
    }
    pivot = pivot.reindex(
        columns=sorted(
            pivot.columns,
            key=lambda item: (group_order.get(str(item[0]), 99), str(item[1])),
        )
    )
    pivot.columns = [f"{group}_{reason}" for group, reason in pivot.columns]
    return summary.merge(pivot.reset_index(), on="일자", how="left").fillna(0)


def _classified_change_frame(start: date, end: date) -> pd.DataFrame:
    rows = []
    for item in NxtChangeStore().list_changes(start, end):
        source_context = source_context_for_change(item)
        rows.append(
            {
                "일자": item.change_date,
                "종목코드": item.stock_code,
                "종목명": item.stock_name,
                "상장시장": item.market,
                "원본변동구분": item.change_type,
                "원본사유": item.reason,
                "재분류구분": classify_nxt_change(item),
                "변경사유": contextual_change_reason(item),
                "데이터근거": source_context.source_title,
                "근거URL": source_context.source_url,
                "보정여부": "보정" if item.is_inferred else "원본",
                "보정근거": item.basis,
            }
        )
    frame = pd.DataFrame(rows)
    if not frame.empty:
        frame["일자"] = pd.to_datetime(frame["일자"])
    return frame


def _adjustment_frame(
    store: HistoricalMarketStore,
    start: date,
    end: date,
) -> pd.DataFrame:
    frame = pd.DataFrame(
        [asdict(item) for item in store.list_nxt_eligibility_adjustments(start, end)]
    ).rename(
        columns={
            "trade_date": "일자",
            "stock_code": "종목코드",
            "stock_name": "종목명",
            "market": "상장시장",
            "unavailable_reason": "거래불가사유",
            "restriction_start_date": "거래불가시작일",
            "restriction_end_date": "거래불가해제일",
            "basis": "보정근거",
        }
    )
    for column in ("일자", "거래불가시작일", "거래불가해제일"):
        if column in frame:
            frame[column] = pd.to_datetime(frame[column])
    return frame


def _classification_rules() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "원본/상태": "변동 원본",
                "원본구분": "편출",
                "사유": "투자경고/위험 지정",
                "최종분류": "거래불가",
                "설명": "매매체결대상 지위는 유지하고 지정기간 거래만 제한",
            },
            {
                "원본/상태": "변동 원본",
                "원본구분": "편출",
                "사유": "단기과열 지정",
                "최종분류": "거래불가",
                "설명": "매매체결대상 지위는 유지하고 지정기간 거래만 제한",
            },
            {
                "원본/상태": "변동 원본",
                "원본구분": "편입",
                "사유": "투자경고/위험 해제",
                "최종분류": "거래불가 해제",
                "설명": "신규 편입이 아니라 기존 대상 종목의 거래제한 해제",
            },
            {
                "원본/상태": "변동 원본",
                "원본구분": "편입",
                "사유": "단기과열 해제",
                "최종분류": "거래불가 해제",
                "설명": "신규 편입이 아니라 기존 대상 종목의 거래제한 해제",
            },
            {
                "원본/상태": "변동 원본",
                "원본구분": "편입",
                "사유": "시가기준가 해제",
                "최종분류": "거래불가 해제",
                "설명": "신규 편입이 아니라 기존 대상 종목의 거래제한 해제",
            },
            {
                "원본/상태": "일별 종목 현황",
                "원본구분": "거래불가",
                "사유": "거래정지·투자경고/위험·단기과열·시가기준가 및 복합사유",
                "최종분류": "거래불가사유",
                "설명": "해당 일자의 NXT 원본 거래불가사유를 그대로 집계",
            },
            {
                "원본/상태": "변동 원본",
                "원본구분": "편입/편출",
                "사유": "위 거래제한 시작·해제 이외의 사유",
                "최종분류": "편입/편출",
                "설명": "실제 매매체결대상 종목 선정의 변동으로 집계",
            },
        ]
    )


def _metadata(start: date, end: date, summary: pd.DataFrame) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ("자료기간", f"{start:%Y-%m-%d}~{end:%Y-%m-%d}"),
            ("수록 거래일", f"{len(summary):,}일"),
            (
                "매매체결대상종목수",
                "NXT 공식 일별 종목 현황에 수록된 고유 종목 수(거래불가 포함)",
            ),
            (
                "거래가능종목수",
                "매매체결대상 중 거래가능시장이 '거래불가'가 아닌 고유 종목 수",
            ),
            (
                "거래불가사유별종목수",
                "NXT 공식 일별 종목 현황의 원본 사유 조합별 고유 종목 수",
            ),
            (
                "편입·편출사유별종목수",
                "거래제한 시작·해제 사유를 제외한 실제 선정 변동 종목 수",
            ),
            (
                "자료원",
                "NXT 정규시장 종목 현황 및 매매체결대상종목 변동내역",
            ),
            (
                "한계",
                "NXT 원본이 투자경고와 투자위험을 합산 표기하므로 둘은 별도 분리하지 않음",
            ),
            (
                "과거 원본 보정",
                "구형 일별 원본에서 빠진 거래제한 종목은 지정·해제 변동과 해제 전 누락기간으로 복원",
            ),
        ],
        columns=["항목", "내용"],
    )


def _validate_frames(
    summaries: pd.DataFrame,
    reasons: pd.DataFrame,
    changes: pd.DataFrame,
) -> None:
    if summaries["일자"].duplicated().any():
        raise RuntimeError("일별요약에 중복 일자가 있습니다.")
    numeric_columns = [column for column in summaries.columns if column != "일자"]
    if summaries[numeric_columns].lt(0).any().any():
        raise RuntimeError("일별요약에 음수 종목수가 있습니다.")
    if not (
        summaries["매매체결대상종목수"]
        == summaries["거래가능종목수"] + summaries["거래불가종목수"]
    ).all():
        raise RuntimeError("매매체결대상 = 거래가능 + 거래불가 관계가 맞지 않습니다.")
    if not (
        summaries["매매체결대상종목수"]
        == summaries["매매체결대상_KOSPI종목수"]
        + summaries["매매체결대상_KOSDAQ종목수"]
    ).all():
        raise RuntimeError("매매체결대상 시장별 합계가 맞지 않습니다.")
    unavailable_by_date = (
        reasons.loc[reasons["구분"] == "거래불가사유"]
        .groupby("일자")["종목수"]
        .sum()
        .reindex(summaries["일자"], fill_value=0)
        .to_numpy()
    )
    if not (unavailable_by_date == summaries["거래불가종목수"].to_numpy()).all():
        raise RuntimeError("거래불가 종목수와 사유별 합계가 맞지 않습니다.")
    if not changes.empty:
        summary_by_date = summaries.set_index("일자")
        count_columns = {
            "편입": "편입종목수",
            "편출": "편출종목수",
            "거래불가": "거래불가지정종목수",
            "거래불가 해제": "거래불가해제종목수",
        }
        for group, column in count_columns.items():
            event_counts = (
                changes.loc[changes["재분류구분"] == group]
                .groupby("일자")["종목코드"]
                .nunique()
                .reindex(summary_by_date.index, fill_value=0)
            )
            if not (event_counts.to_numpy() == summary_by_date[column].to_numpy()).all():
                raise RuntimeError(f"{group} 종목수와 재분류 변동내역이 맞지 않습니다.")
    milestone_expectations = {
        pd.Timestamp("2025-03-24"): (350, 349, 1),
        pd.Timestamp("2025-03-31"): (796, 794, 2),
    }
    summary_by_date = summaries.set_index("일자")
    for milestone, expected in milestone_expectations.items():
        if milestone not in summary_by_date.index:
            continue
        row = summary_by_date.loc[milestone]
        actual = (
            int(row["매매체결대상종목수"]),
            int(row["거래가능종목수"]),
            int(row["거래불가종목수"]),
        )
        if actual != expected:
            raise RuntimeError(
                f"NXT 공식 확대 공지 대사 실패: {milestone:%Y-%m-%d} "
                f"기대 {expected}, 실제 {actual}"
            )
    flow_expectations = {
        pd.Timestamp("2025-03-24"): (240, 0),
        pd.Timestamp("2025-03-31"): (446, 0),
        pd.Timestamp("2025-07-31"): (0, 1),
        pd.Timestamp("2026-01-02"): (223, 152),
    }
    for milestone, expected in flow_expectations.items():
        if milestone not in summary_by_date.index:
            continue
        row = summary_by_date.loc[milestone]
        actual = (int(row["편입종목수"]), int(row["편출종목수"]))
        if actual != expected:
            raise RuntimeError(
                f"공식 선정 명단·일별 대상 명단 대사 실패: {milestone:%Y-%m-%d} "
                f"기대 {expected}, 실제 {actual}"
            )


def _format_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for index, worksheet in enumerate(writer.book.worksheets, start=1):
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if cell.column == 1 and hasattr(cell.value, "year"):
                    cell.number_format = "yyyy-mm-dd"
        for column_cells in worksheet.columns:
            letter = get_column_letter(column_cells[0].column)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[:5000]
            )
            worksheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 45)
        if worksheet.max_row > 1 and worksheet.max_column > 0:
            table = Table(
                displayName=f"NxtHistory{index}",
                ref=worksheet.dimensions,
            )
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            worksheet.add_table(table)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="NXT 일별 매매체결대상·거래가능·사유별 집계를 DB와 Excel로 생성합니다."
    )
    parser.add_argument("--start-date", type=_date_value, default=NXT_LAUNCH_DATE)
    parser.add_argument("--end-date", type=_date_value, default=date.today())
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT
        / "exports"
        / "NXT_일별_매매체결대상_및_사유별_현황.xlsx",
    )
    args = parser.parse_args()
    if args.end_date < args.start_date:
        parser.error("종료일은 시작일보다 빠를 수 없습니다.")

    store = HistoricalMarketStore()
    inferred_count = NxtChangeStore().rebuild_inferred_changes()
    rebuilt_days = store.rebuild_nxt_eligibility_history(
        args.start_date,
        args.end_date,
    )
    summaries = _summary_frame(store, args.start_date, args.end_date)
    reasons = _reason_frame(store, args.start_date, args.end_date)
    if summaries.empty:
        raise RuntimeError("선택한 기간에 저장된 NXT 일별 종목 현황이 없습니다.")
    actual_start = summaries["일자"].min().date()
    actual_end = summaries["일자"].max().date()
    wide_summary = _wide_summary(summaries, reasons)
    changes = _classified_change_frame(actual_start, actual_end)
    adjustments = _adjustment_frame(store, actual_start, actual_end)
    _validate_frames(summaries, reasons, changes)
    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _metadata(actual_start, actual_end, summaries).to_excel(
            writer,
            sheet_name="안내",
            index=False,
        )
        wide_summary.to_excel(writer, sheet_name="일별요약", index=False)
        reasons.to_excel(writer, sheet_name="일별사유집계", index=False)
        for group, sheet_name in (
            ("거래불가사유", "거래불가사유별"),
            ("편출", "편출사유별"),
            ("편입", "편입사유별"),
            ("거래불가", "거래불가지정사유별"),
            ("거래불가 해제", "거래불가해제사유별"),
        ):
            reasons.loc[reasons["구분"] == group, ["일자", "사유", "종목수"]].to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
        changes.to_excel(writer, sheet_name="변동내역_보정포함", index=False)
        adjustments.to_excel(writer, sheet_name="과거거래불가보정", index=False)
        _classification_rules().to_excel(writer, sheet_name="분류기준", index=False)
        _format_workbook(writer)

    coverage = store.nxt_eligibility_history_coverage()
    print(
        f"NXT 일별 선정·거래상태 생성 완료: {rebuilt_days:,}거래일 · "
        f"{coverage['first_date']}~{coverage['last_date']} · "
        f"편입 {coverage['inclusions']:,}건 · 편출 {coverage['exclusions']:,}건 · "
        f"거래불가 지정 {coverage['restrictions']:,}건 · "
        f"거래불가 해제 {coverage['restriction_releases']:,}건"
    )
    print(f"원본 누락 변동 보정: {inferred_count:,}건")
    print(f"Excel: {output_path} ({output_path.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
